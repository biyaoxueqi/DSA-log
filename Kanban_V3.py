# -*- coding: utf-8 -*-
"""
Created on Sat Aug  7 21:38:41 2021

@author: Zhang.Jing
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill, colors, Alignment

from datetime import datetime


Carry_over = []
ECU = []
D = []
MRD_Delivery_Date = []
Delay_days=[]
file_path = 'E:\\Python\\ECUstatus\\SW_Delivery-TestOrder.xlsx'
df = pd.read_excel(file_path)
df=df.fillna(value="")
for i in range(df.shape[0]):
    if df.loc[i,'MRD Delivery Date']=="":
        df.loc[i,'MRD Delivery Date']=df['MRD Date'][0]

for i in range(df.shape[0]):
    if df.loc[i,'SCC Status'] not in ['6. Carry Over', '8. Canceled']:
        Carry_over.append(df.loc[i,'SW Carry Over From'])
        ECU_config = str(df.loc[i,'ECU Instance'])+' '+str(df.loc[i,'ECU Configuration'])
        ECU.append(ECU_config)
        D.append(df.loc[i,'SCC Status'])
        MRD_Delivery_Date.append(df.loc[i,'MRD Delivery Date'])
        
        year=str(20)+df.loc[i,'MRD Delivery Date'][0:2]
        week=df.loc[i,'MRD Delivery Date'][3:5]
        day=df.loc[i,'MRD Delivery Date'][6:7]
        
        datetime_delivery_date=pd.to_datetime(datetime.strptime(year+' '+week+' '+day, '%G %V %u').date())
        datetime_today=pd.to_datetime(datetime.today())
        delay_days=(datetime_today-datetime_delivery_date).days
        if df.loc[i,'SCC Status'] =='1. Planned' and delay_days > 0:
            Delay_days.append(delay_days)
        else:
            Delay_days.append("")
new_df = pd.DataFrame({'Carry_over':Carry_over, 'ECU':ECU, 'SCC_status':D, 'MRD_Delivery_Date':MRD_Delivery_Date,'Delay_days':Delay_days})
list1 = new_df.values.tolist()

column_name= new_df.columns.tolist()

max_row=new_df.shape[0]

interval=25

def area_num(max_row):
    if max_row % interval ==0:
        area_num = int(max_row/20)
    else:
        area_num = int(max_row//20 +1)
    return area_num

k=area_num(max_row)

wb = load_workbook(file_path)
sheet1=wb.create_sheet('Status')

for i in range(k):
    for j in range(len(column_name)):
        sheet1.cell(1,i*len(column_name)+j+1,column_name[j])
    for row_num in range(2,interval+2):
        if i*interval+row_num-2>=len(list1):
            break
        for column_num in range(len(column_name)):
            
            # print(row_num,i*len(column_name)+column_num+1,list1[i*interval+row_num-2][column_num])
            sheet1.cell(row_num,i*len(column_name)+column_num+1,list1[i*interval+row_num-2][column_num])
            
wb.save(file_path)



red_fill = PatternFill("solid", fgColor="FF0000")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
green_fill = PatternFill("solid", fgColor="00FF00")
blue_fill = PatternFill("solid", fgColor="3399FF")
grey_fill = PatternFill("solid", fgColor="C0C0C0")


wb = load_workbook(file_path)
status = wb['Status']
date_example=status['D2'] # 第一行是表头，数据是从第二行开始算的。
today = datetime.now().isocalendar()
today_format = str(today[0])[2:4]+'w'+str(today[1])+str(date_example.value)[5]+str(today[2])

for index, row in enumerate(status.rows):
    if index>0:
        for i in range(k):
            
            cell = row[i*len(column_name)+2]
            delivery_date= row[i*len(column_name)+3]
            if cell.value == '1. Planned' and today_format > delivery_date.value:
                cell.fill = red_fill
            if cell.value == '1. Planned' and today_format <= delivery_date.value:
                cell.fill = yellow_fill
            if cell.value =='4. Under Review':
                cell.fill = blue_fill
            if cell.value =='2. No delivery to Series':
                cell.fill = grey_fill
            if cell.value =='5. OK for Test':
                cell.fill = green_fill
status.insert_rows(1)
wb.save(file_path)

wb = load_workbook(file_path)
status = wb['Status']
#列宽自适应-1/2 生成列名字典，只是为了方便修改列宽时指定列，key:数字，从1开始；value:列名，从A开始
def get_num_colnum_dict():
 # :return: 返回字典：{1:'A', 2:'B', ...... , 52:'AZ'}
    num_str_dict = {}
    A_Z = [chr(a) for a in range(ord('A'), ord('Z') + 1)]
    AA_AZ = ['A' + chr(a) for a in range(ord('A'), ord('Z') + 1)]
    A_AZ = A_Z + AA_AZ
    for i in A_AZ:
        num_str_dict[A_AZ.index(i) + 1] = i
    return num_str_dict

# 列宽自适应-2/2

max_column = status.max_column
max_row = status.max_row
max_column_width_dict = {}
num_str_dict = get_num_colnum_dict()
for i in range(1, max_column + 1):
    for j in range(2, max_row + 1):
        column_width = 0
        cell_value = status.cell(row=j, column=i).value
        cell_value_list = [k for k in str(cell_value)]
        for v in cell_value_list:
            if v.isdigit() == True or v.isalpha() == True:
                column_width += 1
            else:
                column_width += 1.1
        max_column_width_dict.setdefault(i,0)
        if column_width > max_column_width_dict[i]:
            max_column_width_dict[i] = column_width
        """
        try:
            if column_width > max_column_width_dict[i]:
                max_column_width_dict[i] = column_width
        except Exception as e:
                max_column_width_dict[i] = column_width
        """
for key, value in max_column_width_dict.items():
    status.column_dimensions[num_str_dict[key]].width = value

# status.page_setup.fitToHeight status.page_setup.fitToWidth
wb.save(file_path)

# 设置字体
wb = load_workbook(file_path)
status = wb['Status']
fontObj1 = Font(name='微软雅黑', bold=True, italic=False, size=10)
fontObj2 = Font(name='微软雅黑', bold=True, italic=False, size=8)
fontObj3 = Font(name='微软雅黑', bold=False, italic=False, size=8)
status['A1'].font = fontObj1


max_column=status.max_column
max_row=status.max_row
# max_letter=get_column_letter(max_column)

for cellObj in list(status.rows)[1]:
    cellObj.font=fontObj2

for row in range(2, max_row):
    for cellObj in list(status.rows)[row]:
        cellObj.font=fontObj3
wb.save(file_path)

# 设置对齐
wb = load_workbook(file_path)
status = wb['Status']
alignment1 = Alignment(horizontal="center",vertical="center")
alignment2 = Alignment(horizontal="left",vertical="center")
for cellObj in list(status.rows)[0]:
    cellObj.alignment = alignment2
for cellObj in list(status.rows)[1]:
    cellObj.alignment = alignment1
for row in range(2, max_row):
    for cellObj in list(status.rows)[row]:
        cellObj.alignment = alignment2
wb.save(file_path)

# 设置边框


wb = load_workbook(file_path)

status = wb['Status']
max_column = status.max_column
max_row = status.max_row
side1=Side(border_style='thick',color='000000')
border1 = Border(left=side1)
border2 = Border(right=side1)
border3 = Border(top=side1)
border4 = Border(bottom=side1)
border5 = Border(left=side1,bottom=side1)
border6 = Border(left=side1,right=side1,top=side1,bottom=side1)
border7 = Border(right=side1,bottom=side1)


#start='A1'
column_letter=get_column_letter(max_column)
end=column_letter+'1'
# status.merge_cells(start+':'+end)

status['A1']=df.loc[i,'Project Series'] + ' Test order提交进展：'
status['A1'].border=Border(left=side1,top=side1,bottom=side1)
status[end].border=Border(right=side1,top=side1,bottom=side1)
for i in range(2, max_column):
    status[get_column_letter(i)+'1'].border=Border(top=side1,bottom=side1)
"""
for cellObj in list(status.rows)[0]:
    cellObj.border = border6
"""
for cellObj in list(status.rows)[1]:
    cellObj.border = border4
status['A2'].border = border5
status[get_column_letter(max_column)+'2'].border =border7
for cellObj in list(status.rows)[max_row-1]:
    cellObj.border = border4
status['A'+str(max_row)].border=border5
status[get_column_letter(max_column)+str(max_row)].border=border7
for i in range(3,max_row):
    status['A'+str(i)].border =border1
    status[get_column_letter(max_column)+str(i)].border = border2

wb.save(file_path)




today = datetime.now().isocalendar()
today_format = str(today[0])[2:4]+'w'+str(today[1])+str(date_example.value)[5]+str(today[2])
wb = load_workbook(file_path)
status = wb['Status']
max_column = status.max_column
max_row = status.max_row
font_explain = Font(name='微软雅黑', bold=True, italic=False, size=8)
red_fill = PatternFill("solid", fgColor="FF0000")
yellow_fill = PatternFill("solid", fgColor="FFFF00")
green_fill = PatternFill("solid", fgColor="00FF00")
blue_fill = PatternFill("solid", fgColor="3399FF")
grey_fill = PatternFill("solid", fgColor="C0C0C0")
status_list=['SCC审核','已经释放','计划中','延期','不提交']
for i in range(len(status_list)):
    status.cell(row=max_row+2+i,column=max_column-2).value=status_list[i]
    status[get_column_letter(max_column-2)+str(max_row+2+i)].font=font_explain
    status[get_column_letter(max_column)+str(max_row+2+i)].font=font_explain
    if status.cell(row=max_row+2+i,column=max_column-2).value == '延期':
        status.cell(row=max_row+2+i,column=max_column-1).fill = red_fill 
        count_delay=0
        for j in range(3,max_row+1):
            for x in range(k):
                if status.cell(row=j,column=3+x*len(column_name)).value == '1. Planned' and today_format > status.cell(row=j,column=4).value:
                    count_delay+=1
        status.cell(row=max_row+2+i,column=max_column).value =count_delay
        
    if status.cell(row=max_row+2+i,column=max_column-2).value == '计划中':
        status.cell(row=max_row+2+i,column=max_column-1).fill = yellow_fill
        count_planned=0
        for j in range(3,max_row+1): 
            for x in range(k):
                if status.cell(row=j,column=3+x*len(column_name)).value == '1. Planned' and today_format <= status.cell(row=j,column=4).value:
                    count_planned+=1
        status.cell(row=max_row+2+i,column=max_column).value =count_planned
        
    if status.cell(row=max_row+2+i,column=max_column-2).value =='SCC审核':
        status.cell(row=max_row+2+i,column=max_column-1).fill = blue_fill
        count_UnderReview=0  
        for j in range(3,max_row+1):  
            for x in range(k):
                if status.cell(row=j,column=3+x*len(column_name)).value =='4. Under Review':
                    count_UnderReview+=1
        status.cell(row=max_row+2+i,column=max_column).value =count_UnderReview
        
    if status.cell(row=max_row+2+i,column=max_column-2).value =='不提交':
        status.cell(row=max_row+2+i,column=max_column-1).fill = grey_fill
        count_NoDelivery=0  
        for j in range(3,max_row+1):  
            for x in range(k):
                if status.cell(row=j,column=3+x*len(column_name)).value =='2. No delivery to Series':
                    count_NoDelivery+=1
        status.cell(row=max_row+2+i,column=max_column).value =count_NoDelivery
        
    if status.cell(row=max_row+2+i,column=max_column-2).value =='已经释放':
        status.cell(row=max_row+2+i,column=max_column-1).fill = green_fill    
        count_Released=0  
        for j in range(3,max_row+1):  
            for x in range(k):
                if status.cell(row=j,column=3+x*len(column_name)).value =='5. OK for Test':
                    count_Released+=1
        status.cell(row=max_row+2+i,column=max_column).value =count_Released

wb.save(file_path)
